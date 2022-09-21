import openpyxl

# Criar uma Planilha
book = openpyxl.Workbook()
# Como visualizar paginas existentes
print(book.sheetnames)
# Criar SHEETs/Pagina
book.create_sheet('Frutas')
# Como selecionar uma página
frutas_page = book['Frutas']
frutas_page.append(['Banana','5','R$9'])
frutas_page.append(['Banana','5','R$9'])
frutas_page.append(['Banana','5','R$9'])
frutas_page.append(['Banana','5','R$9'])
frutas_page.append(['Banana','5','R$9'])
# Salvar a planilha
book.save('Planilha.xlsx')




